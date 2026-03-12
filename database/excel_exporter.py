"""
Экспорт данных в Excel
Генерация отчетов по прогрессу обучения
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from typing import List, Dict
import os


class ExcelExporter:
    """Класс для экспорта данных в Excel с форматированием"""
    
    def __init__(self):
        self.output_dir = "/home/claude/exports"
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate_progress_report(self, report_data: List[Dict], courses: List[Dict]) -> str:
        """
        Генерирует отчет по прогрессу обучения в Excel
        
        Args:
            report_data: Список данных пользователей с прогрессом
            courses: Список курсов
            
        Returns:
            Путь к созданному файлу
        """
        # Подготовка данных для DataFrame
        rows = []
        
        for user in report_data:
            row = {
                'ID': user['user_id'],
                'Имя': user['full_name'],
                'Дата регистрации': user['registration_date']
            }
            
            # Добавляем колонки для каждого курса
            for course in courses:
                slug = course['course_slug']
                progress = user['courses'].get(slug, {})
                
                completed = progress.get('completed', 0)
                total = progress.get('total', 0)
                is_finished = progress.get('is_finished', False)
                enrollment_date = progress.get('enrollment_date', None)
                
                # Текст статуса
                if enrollment_date:
                    status_text = f"{'✅' if is_finished else '⏸️'} {completed}/{total}"
                else:
                    status_text = "—"
                
                row[course['course_name']] = status_text
                row[f"{course['course_name']}_completed"] = completed
                row[f"{course['course_name']}_total"] = total
                row[f"{course['course_name']}_finished"] = is_finished
                row[f"{course['course_name']}_enrolled"] = enrollment_date is not None
                # Красивое название для даты начала
                row[f"Дата начала ({course['course_name']})"] = enrollment_date or "—"
            
            rows.append(row)
        
        # Создание DataFrame
        df = pd.DataFrame(rows)
        
        # Видимые колонки
        visible_columns = ['ID', 'Имя', 'Дата регистрации']
        for course in courses:
            visible_columns.append(course['course_name'])
            visible_columns.append(f"Дата начала ({course['course_name']})")
        
        df_visible = df[visible_columns]
        
        # Генерация имени файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"progress_report_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Сохранение в Excel
        df_visible.to_excel(filepath, index=False, sheet_name='Прогресс обучения')
        
        # Применение форматирования
        self._apply_formatting(filepath, df, courses)
        
        return filepath
    
    def _apply_formatting(self, filepath: str, df: pd.DataFrame, courses: List[Dict]):
        """Применяет цветовое форматирование к Excel-файлу"""
        
        # Загрузка workbook
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Светло-зеленый
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Желтый
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")    # Серый
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Форматирование заголовков
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Определяем индексы колонок курсов
        course_columns = {}
        for idx, col_name in enumerate(ws[1], start=1):
            col_value = col_name.value
            for course in courses:
                if col_value == course['course_name']:
                    course_columns[course['course_slug']] = idx
                    break
        
        # Форматирование ячеек с прогрессом
        for row_idx in range(2, ws.max_row + 1):
            user_data_idx = row_idx - 2
            
            if user_data_idx >= len(df):
                break
            
            user_row = df.iloc[user_data_idx]
            
            # Форматируем ячейки курсов
            for course in courses:
                slug = course['course_slug']
                col_idx = course_columns.get(slug)
                
                if not col_idx:
                    continue
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Определяем цвет на основе прогресса
                is_enrolled = user_row.get(f"{course['course_name']}_enrolled", False)
                is_finished = user_row.get(f"{course['course_name']}_finished", False)
                completed = user_row.get(f"{course['course_name']}_completed", 0)
                total = user_row.get(f"{course['course_name']}_total", 1)
                
                if not is_enrolled:
                    cell.fill = gray_fill
                elif is_finished:
                    cell.fill = green_fill
                    cell.font = Font(bold=True)
                elif completed > 0:
                    cell.fill = yellow_fill
                else:
                    cell.fill = gray_fill
            
            # Форматирование других ячеек
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                # Центрирование для ID
                if col_idx == 1:
                    cell.alignment = center_alignment
        
        # Автоширина колонок
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Закрепление первой строки
        ws.freeze_panes = 'A2'
        
        # Добавляем лист со статистикой
        self._add_statistics_sheet(wb, df, courses)
        
        # Сохранение
        wb.save(filepath)
    
    def _add_statistics_sheet(self, wb, df: pd.DataFrame, courses: List[Dict]):
        """Добавляет лист со сводной статистикой"""
        
        ws_stats = wb.create_sheet(title="Статистика")
        
        # Заголовок
        ws_stats['A1'] = "Сводная статистика по курсам"
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats.merge_cells('A1:E1')
        
        # Заголовки таблицы
        headers = ['Курс', 'Всего записано', 'Завершили', 'В процессе', '% завершения']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_stats.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        # Данные статистики
        row_idx = 4
        for course in courses:
            enrolled = df[f"{course['course_name']}_enrolled"].sum()
            finished = df[f"{course['course_name']}_finished"].sum()
            in_progress = enrolled - finished
            completion_rate = (finished / enrolled * 100) if enrolled > 0 else 0
            
            ws_stats.cell(row=row_idx, column=1).value = course['course_name']
            ws_stats.cell(row=row_idx, column=2).value = enrolled
            ws_stats.cell(row=row_idx, column=3).value = finished
            ws_stats.cell(row=row_idx, column=4).value = in_progress
            ws_stats.cell(row=row_idx, column=5).value = f"{completion_rate:.1f}%"
            
            row_idx += 1
        
        # Автоширина
        from openpyxl.utils import get_column_letter
        
        for col_idx in range(1, ws_stats.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, ws_stats.max_row + 1):
                cell = ws_stats.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and not isinstance(cell, type(ws_stats.cell(1, 1)).__bases__[0]):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            ws_stats.column_dimensions[column_letter].width = max(max_length + 3, 15)
    
    def generate_user_activity_report(self, user_id: int, progress_data: Dict, 
                                     username: str) -> str:
        """
        Генерирует персональный отчет пользователя
        
        Args:
            user_id: ID пользователя
            progress_data: Словарь с прогрессом по курсам
            username: Имя пользователя
            
        Returns:
            Путь к созданному файлу
        """
        rows = []
        
        for course_slug, progress in progress_data.items():
            course_name = course_slug.replace('_', ' ').title()
            completed = progress.get('completed', 0)
            total = progress.get('total', 0)
            percentage = progress.get('percentage', 0)
            
            rows.append({
                'Курс': course_name,
                'Пройдено уроков': f"{completed}/{total}",
                'Процент завершения': f"{percentage}%",
                'Статус': '✅ Завершен' if progress.get('is_finished') else '⏸️ В процессе'
            })
        
        df = pd.DataFrame(rows)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"user_{user_id}_activity_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        df.to_excel(filepath, index=False, sheet_name=f'Активность {username}')
        
        return filepath


# Глобальный экземпляр
exporter = ExcelExporter()